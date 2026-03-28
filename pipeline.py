"""
Bridge – Data Pipeline
Converts provider Excel/CSV files into the JSON format the frontend expects.

Usage:
    python pipeline.py --provider quilter --input ./data/ --output ./output/

The script reads Excel files from the input directory, transforms them into
the D= blob structure, and writes the JSON to the output directory.

To integrate into index.html:
    1. Run this script
    2. Open the generated JSON
    3. Replace the const D={...}; blob in index.html with the new data

Or use --inject flag to automatically inject into an existing index.html:
    python pipeline.py --provider quilter --input ./data/ --inject ./index.html
"""

import openpyxl
import json
import os
import sys
import argparse
from datetime import datetime


# ─── Excel Reader ──────────────────────────────────────────────────────

def read_sheet(path, sheet=None):
    """Read an Excel sheet into a list of dicts."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]
    data = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        data.append(dict(zip(headers, row)))
    return data


def ser(v):
    """Serialize datetime objects to ISO strings."""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return v


def clean(rows):
    """Clean a list of dicts, serializing datetimes."""
    return [{k: ser(v) for k, v in r.items()} for r in rows]


# ─── Quilter WealthSelect Managed Active ──────────────────────────────

def build_quilter(input_dir):
    """Build the full data blob from Quilter Excel files."""

    def load(filename, sheet=None):
        path = os.path.join(input_dir, filename)
        if not os.path.exists(path):
            print(f"  ⚠ Missing: {filename}")
            return []
        return read_sheet(path, sheet)

    # ─── Load template-based data ─────────────────────────────────────
    # The main template file contains Historical Holdings + other tabs
    template_file = None
    for f in os.listdir(input_dir):
        if 'Historical_Holdings' in f and f.endswith('.xlsx'):
            template_file = f
            break
        if 'Template' in f and f.endswith('.xlsx'):
            template_file = f
            break

    if template_file:
        print(f"  ✓ Using template: {template_file}")
        hist_holdings = load(template_file, "Historical Holdings")
    else:
        print(f"  ⚠ No template file found, trying individual files")
        hist_holdings = []

    # Load other data sources
    cal_all = load("Quilter_WS_Active_-_Calendar_Returns.xlsx")
    trail_all = load("Quilter_WS_Active_-_Trailing_Returns.xlsx")
    additional = load("Additional_Content_for_Bridge.xlsx", "Text")
    risk_return = load("Additional_Content_for_Bridge.xlsx", "Risk Return")

    # Parse additional content into lookup
    content = {}
    for r in additional:
        sec = r.get("Section", "")
        sub = r.get("Sub Section", "")
        txt = r.get("Content", "")
        if sec not in content:
            content[sec] = {}
        content[sec][sub] = txt

    ip = content.get("Investment Process", {})
    adv = content.get("Adviser Support", {})
    qr = content.get("Quarterly Review", {})

    # Build risk/return data for provider
    rr_data = []
    for r in risk_return:
        portfolio = r.get("Portfolio", "")
        period = r.get("Time Period", "")
        risk_val = r.get("Risk (Annualised)", 0)
        ret_val = r.get("Return (Annualised)", 0)
        if portfolio and period:
            rr_data.append({
                "period": period,
                "portfolio": portfolio,
                "risk": round(risk_val, 2) if risk_val else 0,
                "return": round(ret_val, 2) if ret_val else 0,
            })

    # Separate Quilter rows from EAA benchmarks
    cal = [r for r in cal_all if r.get("Model", "").startswith("Quilter")]
    trail = [r for r in trail_all if r.get("Model", "").startswith("Quilter")]
    benchmarks_trail = [r for r in trail_all if r.get("Model", "").startswith("EAA")]
    benchmarks_cal = [r for r in cal_all if r.get("Model", "").startswith("EAA")]

    # ─── Process Historical Holdings ──────────────────────────────────
    # Parse dates and extract risk levels
    from collections import defaultdict

    def parse_date(d):
        if isinstance(d, datetime):
            return d
        if isinstance(d, str):
            for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y'):
                try:
                    return datetime.strptime(d.split(' ')[0], fmt)
                except ValueError:
                    continue
        return None

    def extract_risk_level(portfolio_name):
        if not portfolio_name:
            return None
        parts = portfolio_name.strip().split()
        try:
            return int(parts[-1])
        except (ValueError, IndexError):
            return None

    # Group holdings by date and risk level
    holdings_by_date_risk = defaultdict(lambda: defaultdict(list))
    all_dates = set()

    for h in hist_holdings:
        dt = parse_date(h.get("Date"))
        portfolio = h.get("Portfolio", "")
        risk_level = extract_risk_level(portfolio)
        if dt is None or risk_level is None:
            continue
        all_dates.add(dt)
        holdings_by_date_risk[dt][risk_level].append({
            "name": h.get("Fund Name", ""),
            "isin": h.get("ISIN", ""),
            "weight": round(float(h.get("Weight (%)", 0) or 0), 4),
            "type": h.get("Asset Class", ""),
            "sub_type": h.get("Sub Asset Class", ""),
        })

    sorted_dates = sorted(all_dates)
    latest_date = sorted_dates[-1] if sorted_dates else None
    risk_levels_found = sorted(set(rl for d in holdings_by_date_risk for rl in holdings_by_date_risk[d]))

    print(f"  Historical Holdings: {len(hist_holdings)} rows")
    print(f"  Dates: {len(sorted_dates)} ({sorted_dates[0].strftime('%d/%m/%Y') if sorted_dates else 'none'} to {sorted_dates[-1].strftime('%d/%m/%Y') if sorted_dates else 'none'})")
    print(f"  Risk levels: {risk_levels_found}")
    print(f"  Latest date: {latest_date.strftime('%d/%m/%Y') if latest_date else 'none'}")

    # ─── Current exposure (from latest date) ──────────────────────────
    def compute_exposure(holdings_list):
        """Compute asset allocation, geo, FI breakdown from a list of holdings."""
        aa = {"equity": 0, "bonds": 0, "alternatives": 0, "cash": 0}
        geo = defaultdict(float)
        fi_breakdown = defaultdict(float)

        for h in holdings_list:
            w = h["weight"]
            ac = h["type"]
            sc = h["sub_type"]

            if ac == "Equity":
                aa["equity"] += w
                if sc:
                    geo[sc.lower().replace(" ", "_")] += w
            elif ac == "Fixed Income":
                aa["bonds"] += w
                if sc:
                    fi_breakdown[sc.lower().replace(" ", "_")] += w
            elif ac == "Alternative":
                aa["alternatives"] += w
            elif ac == "Cash":
                aa["cash"] += w

        aa = {k: round(v, 2) for k, v in aa.items()}
        geo = {k: round(v, 4) for k, v in geo.items()}
        fi_breakdown = {k: round(v, 4) for k, v in fi_breakdown.items()}
        return aa, geo, fi_breakdown

    # Cost data
    cost_data = {
        3:  {"dfm": 0.15, "underlying": 0.45, "all_in": 0.60},
        4:  {"dfm": 0.15, "underlying": 0.52, "all_in": 0.67},
        5:  {"dfm": 0.15, "underlying": 0.58, "all_in": 0.73},
        6:  {"dfm": 0.15, "underlying": 0.63, "all_in": 0.78},
        7:  {"dfm": 0.15, "underlying": 0.66, "all_in": 0.81},
        8:  {"dfm": 0.15, "underlying": 0.68, "all_in": 0.83},
        9:  {"dfm": 0.15, "underlying": 0.66, "all_in": 0.81},
        10: {"dfm": 0.15, "underlying": 0.64, "all_in": 0.79},
    }

    platforms = [
        "AJ Bell Investcentre", "Fidelity Adviser Solutions", "M&G Wealth",
        "Morningstar Wealth", "Parmenion", "Quilter", "Transact"
    ]

    # Provider metadata
    provider = {
        "id": "quilter",
        "name": "Quilter Investors",
        "full_name": "Quilter WealthSelect Managed Active",
        "investment_style": "Active",
        "description": "",
        "aum_bn": 0,
        "established": "",
        "headquarters": "",
        "key_personnel": [],
        "strengths": [],
        "considerations": [],
        "regulatory_status": "",
        "website": "https://www.quilter.com/investments/wealthselect-managed-portfolios/",
        "risk_rating_providers": ["Dynamic Planner", "EV", "Finametrica", "Synaptic"],
        "investment_process": {
            "investment_team": ip.get("Investment Team", ""),
            "strategic_aa": ip.get("Strategic Asset Allocation", ""),
            "tactical_aa": ip.get("Tactical Asset Allocation", ""),
            "fund_selection": ip.get("Fund Selection", ""),
            "portfolio_construction": ip.get("Portfolio Construction", ""),
            "rebalancing": ip.get("Rebalancing & Trading", "")
        },
        "adviser_support": {
            "onboarding": adv.get("Onboarding & Training", ""),
            "ongoing_communication": adv.get("Ongoing Portfolio Communication", ""),
            "review_meetings": adv.get("Review Meetings & Ongoing Engagement", ""),
            "investment_commentary": adv.get("Investment Commentary & Market Insight", ""),
            "digital_tools": adv.get("Digital Tools & Support Infrastructure", "")
        },
        "quarterly_review": {
            "aa_changes": qr.get("Asset Allocation Changes", ""),
            "fund_changes": qr.get("Fund Selection Changes", ""),
            "performance_review": qr.get("Performance Review", "")
        },
        "risk_return_data": rr_data,
    }

    # ─── Build MPS models ─────────────────────────────────────────────
    models = []
    for rl in risk_levels_found:
        # Find trailing returns row
        trail_name = f"Quilter WealthSelect Active Managed {rl}"
        t = next((r for r in trail if r.get("Model") == trail_name), {})

        # Current holdings from latest date
        current_holdings = holdings_by_date_risk.get(latest_date, {}).get(rl, [])
        aa, geo, fi_breakdown = compute_exposure(current_holdings)

        # Calendar returns
        cal_row = next((c for c in cal if c.get("Model") == trail_name), {})
        calendar = {}
        for k, v in cal_row.items():
            if k != "Model" and v is not None:
                calendar[str(k)] = v

        costs = cost_data.get(rl, {"dfm": 0.15, "underlying": 0.60, "all_in": 0.75})

        # Sort holdings by weight descending
        uf = sorted(current_holdings, key=lambda x: x.get("weight", 0), reverse=True)

        model = {
            "id": f"quilter-ws-active-{rl}",
            "name": trail_name if t else f"Quilter WealthSelect Active Managed {rl}",
            "provider": "Quilter Investors",
            "risk_rating": rl,
            "risk_label": f"Risk {rl}",
            "asset_allocation": aa,
            "geographic_allocation": geo,
            "fixed_income_breakdown": fi_breakdown,
            "ocf": costs["all_in"],
            "cost_breakdown": costs,
            "return_1yr": t.get("1Y"),
            "return_3yr": t.get("3Y"),
            "return_5yr": t.get("5Y"),
            "return_10yr": t.get("10Y"),
            "return_ytd": t.get("YTD"),
            "return_1m": t.get("1M"),
            "return_3m": t.get("3M"),
            "return_6m": t.get("6M"),
            "calendar_returns": calendar,
            "trailing_returns": {k: v for k, v in t.items() if k != "Model"},
            "rebalancing": "Quarterly",
            "min_investment": 0,
            "platforms": platforms,
            "ethical": False,
            "decumulation_suitable": rl <= 5,
            "time_horizons": (
                ["short", "medium", "long"] if rl <= 4
                else ["medium", "long"] if rl <= 6
                else ["long"]
            ),
            "underlying_funds": uf,
            "inception_date": "2014-02-24",
        }
        models.append(model)

    # ─── Build historical data for ALL risk levels ────────────────────
    historical = {}
    for rl in risk_levels_found:
        hist_aa = []  # asset allocation over time
        hist_eq = []  # equity region over time
        hist_fi = []  # fixed income over time

        for dt in sorted_dates:
            holdings = holdings_by_date_risk.get(dt, {}).get(rl, [])
            if not holdings:
                continue

            date_str = dt.strftime("%Y-%m-%d")
            aa, geo, fi_breakdown = compute_exposure(holdings)

            # Asset allocation row
            aa_row = {"Date": date_str}
            aa_row.update(aa)
            hist_aa.append(aa_row)

            # Equity region row
            eq_row = {"Date": date_str}
            eq_row.update(geo)
            hist_eq.append(eq_row)

            # Fixed income row
            fi_row = {"Date": date_str}
            fi_row.update(fi_breakdown)
            hist_fi.append(fi_row)

        historical[f"portfolio_{rl}"] = {
            "asset_allocation": hist_aa,
            "equity_region": hist_eq,
            "fixed_income": hist_fi,
        }
        print(f"  Historical portfolio_{rl}: {len(hist_aa)} dates")

    # ─── Assemble output ──────────────────────────────────────────────
    output = {
        "mps": models,
        "providers": {"Quilter Investors": provider},
        "platforms": platforms,
        "styles": ["Active"],
        "insights": [],
        "performance": {},
        "historical": historical,
        "benchmarks": {
            "trailing": clean(benchmarks_trail),
            "calendar": clean(benchmarks_cal),
        },
        "cost_table": [{"model": m["name"], **m["cost_breakdown"]} for m in models],
    }

    return output, models


# ─── Inject into index.html ──────────────────────────────────────────

def inject_into_html(data, html_path, output_path=None):
    """Replace the const D={...}; blob in index.html with new data."""
    with open(html_path, "r") as f:
        html = f.read()

    # Find and replace the data blob
    start_marker = "const D="
    end_marker = ";\n"

    start_idx = html.index(start_marker)
    # Find the semicolon+newline after the JSON blob
    # The blob is a single line of JSON, so find the next ;\n after start
    json_start = start_idx + len(start_marker)
    # Find matching end - the JSON is one line so look for ;\n
    search_from = json_start
    depth = 0
    for i in range(search_from, len(html)):
        if html[i] == "{":
            depth += 1
        elif html[i] == "}":
            depth -= 1
            if depth == 0:
                end_idx = i + 1
                break

    old_blob = html[start_idx:end_idx + 1]  # includes the ;
    new_blob = f"const D={json.dumps(data, default=str)};"

    html = html[:start_idx] + new_blob + html[end_idx + 1:]

    out = output_path or html_path
    with open(out, "w") as f:
        f.write(html)

    print(f"✓ Injected into {out}")
    print(f"  Old blob: {len(old_blob)} chars")
    print(f"  New blob: {len(new_blob)} chars")


# ─── CLI ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Bridge Data Pipeline")
    parser.add_argument("--provider", default="quilter", help="Provider name (default: quilter)")
    parser.add_argument("--input", default="./data/", help="Input directory with Excel files")
    parser.add_argument("--output", default="./output/", help="Output directory for JSON")
    parser.add_argument("--inject", default=None, help="Path to index.html to inject data into")
    args = parser.parse_args()

    print(f"Bridge Data Pipeline")
    print(f"Provider: {args.provider}")
    print(f"Input: {args.input}")
    print()

    if args.provider == "quilter":
        data, models = build_quilter(args.input)
    else:
        print(f"Unknown provider: {args.provider}")
        print("Supported: quilter")
        sys.exit(1)

    # Print summary
    print(f"\n{'='*50}")
    print(f"SUMMARY")
    print(f"{'='*50}")
    print(f"Portfolios: {len(models)}")
    for m in models:
        aa = m["asset_allocation"]
        print(f"  {m['name']}")
        print(f"    Eq {aa['equity']}% | FI {aa['bonds']}% | Alt {aa['alternatives']}% | Cash {aa['cash']}%")
        print(f"    OCF {m['ocf']}% | 1Y {m['return_1yr']}% | 3Y {m['return_3yr']}% | 5Y {m['return_5yr']}%")
    hist = data.get("historical", {})
    for key in sorted(hist.keys()):
        h = hist[key]
        print(f"Historical {key}: AA={len(h.get('asset_allocation', []))} Eq={len(h.get('equity_region', []))} FI={len(h.get('fixed_income', []))}")
    print(f"Total fund holdings: {sum(len(m['underlying_funds']) for m in models)}")
    print(f"Benchmark rows: {len(data.get('benchmarks', {}).get('trailing', []))}")

    # Write JSON
    os.makedirs(args.output, exist_ok=True)
    json_path = os.path.join(args.output, f"{args.provider}_data.json")
    with open(json_path, "w") as f:
        json.dump(data, f, indent=2, default=str)
    print(f"\n✓ JSON saved to {json_path}")

    # Inject if requested
    if args.inject:
        inject_into_html(data, args.inject)

    print("\nDone.")


if __name__ == "__main__":
    main()
